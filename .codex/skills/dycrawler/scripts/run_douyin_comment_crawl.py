#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import subprocess
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any


LOCAL_FALLBACK_REPO = Path("/Users/lulu/AIWork/MediaCrawler")
CHINA_TZ = timezone(timedelta(hours=8))


def parse_aweme_id(value: str | None) -> str:
    if not value:
        return ""
    patterns = [
        r"modal_id=(\d+)",
        r"/video/(\d+)",
        r"aweme_id=(\d+)",
        r"(^|\D)(\d{16,22})(\D|$)",
    ]
    for pattern in patterns:
        match = re.search(pattern, value)
        if not match:
            continue
        groups = [g for g in match.groups() if g and g.isdigit()]
        if groups:
            return groups[0]
    stripped = value.strip()
    return stripped if stripped.isdigit() else ""


def stable_slug(source: str, aweme_id: str) -> str:
    if aweme_id:
        return aweme_id
    digest = hashlib.sha1(source.encode("utf-8")).hexdigest()[:12]
    return f"unknown_{digest}"


def is_mediacrawler_repo(path: Path) -> bool:
    return (path / "main.py").exists() and (path / "pyproject.toml").exists()


def discover_mediacrawler_dir(cli_value: str) -> Path:
    if cli_value:
        return Path(cli_value).expanduser().resolve()

    env_value = os.environ.get("MEDIACRAWLER_DIR", "")
    if env_value:
        return Path(env_value).expanduser().resolve()

    search_roots = [Path.cwd(), *Path.cwd().parents, Path(__file__).resolve().parent]
    for root in search_roots:
        candidates = [root, root / "MediaCrawler"]
        for candidate in candidates:
            if is_mediacrawler_repo(candidate):
                return candidate.resolve()

    return LOCAL_FALLBACK_REPO.resolve()


def require_repo(repo: Path) -> None:
    missing = [path for path in [repo / "main.py", repo / "pyproject.toml"] if not path.exists()]
    if missing:
        raise SystemExit(f"MediaCrawler repo is invalid or missing files: {missing}")


def check_required_patch(repo: Path, allow_unpatched: bool) -> None:
    core_path = repo / "media_platform/douyin/core.py"
    client_path = repo / "media_platform/douyin/client.py"
    core = core_path.read_text(encoding="utf-8")
    client = client_path.read_text(encoding="utf-8")
    checks = {
        "core uses asyncio.gather": "await asyncio.gather(*task_list)" in core,
        "core re-raises unexpected errors": "unexpected comment crawl error" in core and "raise" in core,
        "client tracks top-level comments": "top_level_count" in client,
        "client tracks sub-comments": "sub_comment_count" in client,
        "client warns on empty pages": "empty top-level comments page" in client and "empty sub-comments page" in client,
        "client logs final total": "finished, " in client and "top_level_count" in client,
    }
    failed = [name for name, ok in checks.items() if not ok]
    if not failed:
        return
    message = "MediaCrawler Douyin reliability patch is missing: " + "; ".join(failed)
    if allow_unpatched:
        print(f"[WARN] {message}", file=sys.stderr)
        return
    raise SystemExit(message)


def run_mediacrawler(
    repo: Path,
    source: str,
    out_root: Path,
    sleep_sec: float,
    max_comments: int,
    include_sub_comments: bool,
    login_type: str,
    headless: bool,
) -> None:
    out_root.mkdir(parents=True, exist_ok=True)
    inline = f"""
import asyncio
import sys

import config
import main as app_main

config.CRAWLER_MAX_SLEEP_SEC = {sleep_sec!r}
sys.argv = [
    "main.py",
    "--platform", "dy",
    "--lt", {login_type!r},
    "--type", "detail",
    "--specified_id", {source!r},
    "--get_comment", "true",
    "--get_sub_comment", {"true" if include_sub_comments else "false"!r},
    "--max_comments_count_singlenotes", {str(max_comments)!r},
    "--max_concurrency_num", "1",
    "--headless", {"true" if headless else "false"!r},
    "--save_data_option", "jsonl",
    "--save_data_path", {str(out_root)!r},
]
asyncio.run(app_main.main())
"""
    cmd = ["uv", "run", "python", "-c", inline]
    proc = subprocess.Popen(cmd, cwd=repo)
    return_code = proc.wait()
    if return_code != 0:
        raise SystemExit(f"MediaCrawler failed with exit code {return_code}")


def find_latest_comments_jsonl(out_root: Path) -> Path:
    candidates = sorted(
        (out_root / "douyin" / "jsonl").glob("detail_comments_*.jsonl"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise SystemExit(f"No detail_comments_*.jsonl found under {out_root}")
    return candidates[0]


def is_top_level(row: dict[str, Any]) -> bool:
    return str(row.get("parent_comment_id", "0")) in {"0", "None", ""}


def text(value: Any) -> str:
    return "" if value is None else str(value)


def created_at(value: Any) -> str:
    try:
        return datetime.fromtimestamp(int(value), tz=CHINA_TZ).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, 1):
            if not line.strip():
                continue
            row = json.loads(line)
            row["_crawl_order"] = line_number
            rows.append(row)
    if not rows:
        raise SystemExit(f"No comments found in {path}")
    return rows


def export_tables(rows: list[dict[str, Any]], out_root: Path, source_url: str, fallback_aweme_id: str) -> dict[str, Any]:
    parent_map = {text(row.get("comment_id")): row for row in rows}
    aweme_id = text(rows[0].get("aweme_id")) or fallback_aweme_id or "unknown"
    export_rows: list[dict[str, Any]] = []

    for row in rows:
        parent_id = text(row.get("parent_comment_id"))
        parent = parent_map.get(parent_id)
        level = 1 if is_top_level(row) else 2
        export_rows.append(
            {
                "crawl_order": row["_crawl_order"],
                "level": level,
                "comment_id": text(row.get("comment_id")),
                "parent_comment_id": "0" if level == 1 else parent_id,
                "parent_nickname": "" if level == 1 or not parent else text(parent.get("nickname")),
                "parent_content": "" if level == 1 or not parent else text(parent.get("content")),
                "nickname": text(row.get("nickname")),
                "content": text(row.get("content")),
                "like_count": row.get("like_count") or 0,
                "sub_comment_count": row.get("sub_comment_count") or "",
                "ip_location": text(row.get("ip_location")),
                "created_at": created_at(row.get("create_time")),
                "create_time": text(row.get("create_time")),
                "user_unique_id": text(row.get("user_unique_id")),
                "short_user_id": text(row.get("short_user_id")),
                "user_id": text(row.get("user_id")),
                "sec_uid": text(row.get("sec_uid")),
                "avatar": text(row.get("avatar")),
            }
        )

    headers = list(export_rows[0].keys())
    csv_path = out_root / f"douyin_comments_{aweme_id}.csv"
    xlsx_path = out_root / f"douyin_comments_{aweme_id}.xlsx"
    summary_path = out_root / f"douyin_comments_{aweme_id}_summary.json"

    out_root.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(export_rows)

    xlsx_written = False
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "summary"
        top_level_count = sum(1 for row in export_rows if row["level"] == 1)
        sub_comment_count = len(export_rows) - top_level_count
        for item in [
            ["source_url", source_url],
            ["aweme_id", aweme_id],
            ["total_comments", len(export_rows)],
            ["top_level_comments", top_level_count],
            ["sub_comments", sub_comment_count],
            ["exported_at", datetime.now(CHINA_TZ).strftime("%Y-%m-%d %H:%M:%S")],
        ]:
            summary_sheet.append(item)
        summary_sheet.column_dimensions["A"].width = 24
        summary_sheet.column_dimensions["B"].width = 120

        header_fill = PatternFill("solid", fgColor="D9EAF7")

        def add_sheet(name: str, data: list[dict[str, Any]]) -> None:
            sheet = workbook.create_sheet(name)
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for item in data:
                sheet.append([item[header] for header in headers])
            sheet.freeze_panes = "A2"
            widths = {
                "crawl_order": 12,
                "level": 8,
                "comment_id": 24,
                "parent_comment_id": 24,
                "parent_nickname": 18,
                "parent_content": 48,
                "nickname": 18,
                "content": 72,
                "like_count": 12,
                "sub_comment_count": 16,
                "ip_location": 12,
                "created_at": 20,
                "create_time": 14,
                "user_unique_id": 22,
                "short_user_id": 18,
                "user_id": 22,
                "sec_uid": 48,
                "avatar": 48,
            }
            for index, header in enumerate(headers, 1):
                column = get_column_letter(index)
                sheet.column_dimensions[column].width = widths.get(header, 16)
                if header in {"content", "parent_content"}:
                    for cell in sheet[column][1:]:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                if header in {"comment_id", "parent_comment_id", "user_unique_id", "short_user_id", "user_id", "sec_uid"}:
                    for cell in sheet[column][1:]:
                        cell.number_format = "@"

        add_sheet("all_comments", export_rows)
        add_sheet("top_level", [row for row in export_rows if row["level"] == 1])
        add_sheet("sub_comments", [row for row in export_rows if row["level"] == 2])
        workbook.save(xlsx_path)
        xlsx_written = True
    except ImportError:
        xlsx_path = Path("")

    top_level_count = sum(1 for row in export_rows if row["level"] == 1)
    sub_comment_count = len(export_rows) - top_level_count
    summary = {
        "ok": True,
        "source_url": source_url,
        "aweme_id": aweme_id,
        "total_comments": len(export_rows),
        "top_level_comments": top_level_count,
        "sub_comments": sub_comment_count,
        "unique_comment_ids": len({row["comment_id"] for row in export_rows}),
        "csv_path": str(csv_path),
        "xlsx_path": str(xlsx_path) if xlsx_written else None,
        "summary_path": str(summary_path),
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run MediaCrawler Douyin single-post comment crawl and export tables.")
    parser.add_argument("source", nargs="?", help="Douyin post URL, modal_id URL, /video URL, short URL, or aweme ID.")
    parser.add_argument("--mediacrawler-dir", default="", help="Path to the MediaCrawler repo. Defaults to auto-discovery or MEDIACRAWLER_DIR.")
    parser.add_argument("--out-root", default="", help="Output root. Defaults to <MediaCrawler>/output/douyin_comments_<id>_<timestamp>.")
    parser.add_argument("--sleep-sec", type=float, default=4.0, help="Sleep interval between comment API requests.")
    parser.add_argument("--max-comments", type=int, default=99999, help="Maximum top-level comments to request from MediaCrawler.")
    parser.add_argument("--include-sub-comments", dest="include_sub_comments", action="store_true", default=True)
    parser.add_argument("--no-include-sub-comments", dest="include_sub_comments", action="store_false")
    parser.add_argument("--login-type", choices=["qrcode", "phone", "cookie"], default="qrcode")
    parser.add_argument("--headless", action="store_true", help="Run browser headless when MediaCrawler supports it.")
    parser.add_argument("--allow-unpatched", action="store_true", help="Warn instead of aborting if required MediaCrawler patch is missing.")
    parser.add_argument("--export-existing-jsonl", default="", help="Skip crawling and export an existing detail_comments_*.jsonl file.")
    parser.add_argument("--source-url", default="", help="Source URL metadata when using --export-existing-jsonl.")
    parser.add_argument("--aweme-id", default="", help="Fallback aweme ID when exporting an existing JSONL file.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    repo = discover_mediacrawler_dir(args.mediacrawler_dir)

    if args.export_existing_jsonl:
        jsonl_path = Path(args.export_existing_jsonl).expanduser().resolve()
        if not jsonl_path.exists():
            raise SystemExit(f"JSONL file not found: {jsonl_path}")
        source = args.source_url or args.source or ""
        fallback_aweme_id = args.aweme_id or parse_aweme_id(source)
        out_root = Path(args.out_root).expanduser().resolve() if args.out_root else jsonl_path.parents[2]
        rows = load_jsonl(jsonl_path)
        summary = export_tables(rows, out_root, source, fallback_aweme_id)
        summary["jsonl_path"] = str(jsonl_path)
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return

    if not args.source:
        raise SystemExit("A Douyin post URL or aweme ID is required.")

    require_repo(repo)
    check_required_patch(repo, args.allow_unpatched)

    aweme_id = parse_aweme_id(args.source)
    slug = stable_slug(args.source, aweme_id)
    timestamp = datetime.now(CHINA_TZ).strftime("%Y%m%d_%H%M%S")
    out_root = Path(args.out_root).expanduser().resolve() if args.out_root else repo / "output" / f"douyin_comments_{slug}_{timestamp}"

    run_mediacrawler(
        repo=repo,
        source=args.source,
        out_root=out_root,
        sleep_sec=args.sleep_sec,
        max_comments=args.max_comments,
        include_sub_comments=args.include_sub_comments,
        login_type=args.login_type,
        headless=args.headless,
    )
    jsonl_path = find_latest_comments_jsonl(out_root)
    rows = load_jsonl(jsonl_path)
    summary = export_tables(rows, out_root, args.source, aweme_id)
    summary["jsonl_path"] = str(jsonl_path)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
